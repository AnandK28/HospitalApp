"""
db.py — local SQLite data layer for the Hospital Records app.

Mirrors your Postgres schema (patients + hospital_stays) but runs
entirely on-device, no server required.
"""

import sqlite3
import os
from datetime import date

from kivymd.app import MDApp


def get_db_path():
    """Store the DB inside the app's private data dir (works on Android + desktop)."""
    app = MDApp.get_running_app()
    base = app.user_data_dir if app else "."
    return os.path.join(base, "hospital.db")


import json


def get_lock_path():
    app = MDApp.get_running_app()
    base = app.user_data_dir if app else "."
    return os.path.join(base, "lock.json")


def get_lock():
    """Returns {'type': 'pin'|'pattern', 'value': str}. Default PIN 1234 on first run."""
    path = get_lock_path()
    if not os.path.exists(path):
        default = {"type": "pin", "value": "1234"}
        with open(path, "w") as f:
            json.dump(default, f)
        return default
    with open(path) as f:
        return json.load(f)


def set_lock(lock_type, value):
    with open(get_lock_path(), "w") as f:
        json.dump({"type": lock_type, "value": value}, f)


def get_export_dir():
    app = MDApp.get_running_app()
    base = app.user_data_dir if app else "."
    export_dir = os.path.join(base, "exports")
    os.makedirs(export_dir, exist_ok=True)
    return export_dir


def _connect():
    conn = sqlite3.connect(get_db_path())
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = _connect()
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS patients (
            patient_id  INTEGER PRIMARY KEY AUTOINCREMENT,
            first_name  TEXT NOT NULL,
            last_name   TEXT NOT NULL,
            dob         TEXT NOT NULL,
            mrn         TEXT UNIQUE NOT NULL
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS hospital_stays (
            stay_id             INTEGER PRIMARY KEY AUTOINCREMENT,
            patient_id          INTEGER REFERENCES patients(patient_id),
            admission_date      TEXT DEFAULT CURRENT_TIMESTAMP,
            discharge_date      TEXT,
            primary_diagnosis   TEXT NOT NULL,
            daily_progress_notes TEXT,
            hospital_course     TEXT,
            treatment_given     TEXT,
            discharge_condition TEXT,
            discharge_advice    TEXT
        )
    """)

    conn.commit()

    # seed a little demo data on first run only
    cur.execute("SELECT COUNT(*) AS c FROM patients")
    if cur.fetchone()["c"] == 0:
        _seed_demo_data(conn)

    conn.close()


def _seed_demo_data(conn):
    patients = [
        ("Anjali", "Krishnan", "1988-04-12", "MRN-1001"),
        ("Ravi", "Subramaniam", "1975-11-02", "MRN-1002"),
        ("Meera", "Natarajan", "1993-07-21", "MRN-1003"),
    ]
    cur = conn.cursor()
    cur.executemany(
        "INSERT INTO patients (first_name, last_name, dob, mrn) VALUES (?, ?, ?, ?)",
        patients,
    )

    stays = [
        (1, "2026-06-01 09:00", "2026-06-05 14:00", "Community-acquired pneumonia",
         "Day1: febrile, started IV antibiotics.\nDay2: improving, O2 weaned.",
         "Admitted with fever and cough, treated with IV ceftriaxone, responded well.",
         "IV ceftriaxone 1g OD, supportive O2 therapy",
         "Afebrile, ambulatory, saturating well on room air",
         "Complete oral antibiotics, review in OPD after 1 week, return if fever recurs"),
        (2, "2026-06-10 11:00", "2026-06-12 10:00", "Acute gastroenteritis",
         "Day1: dehydrated, IV fluids started.",
         "Presented with vomiting and diarrhea, rehydrated, symptoms resolved.",
         "IV fluids, ondansetron, oral rehydration",
         "Stable, tolerating oral fluids",
         "Continue ORS, bland diet, follow up if symptoms recur"),
        (3, "2026-06-20 08:30", None, "Type 2 diabetes mellitus - uncontrolled",
         "Day1: blood sugar 340, insulin sliding scale started.",
         "Admitted for glycemic control, ongoing insulin titration.",
         "Insulin sliding scale, dietary counseling",
         "Under observation",
         "Pending discharge"),
    ]
    cur.executemany("""
        INSERT INTO hospital_stays
        (patient_id, admission_date, discharge_date, primary_diagnosis,
         daily_progress_notes, hospital_course, treatment_given,
         discharge_condition, discharge_advice)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, stays)
    conn.commit()


def search_patients(query):
    """Equivalent of the ILIKE search across first_name, last_name, primary_diagnosis."""
    conn = _connect()
    like = f"%{query}%"
    rows = conn.execute("""
        SELECT p.patient_id, p.first_name, p.last_name, p.mrn, p.dob,
               s.stay_id, s.primary_diagnosis, s.admission_date, s.discharge_date
        FROM patients p
        JOIN hospital_stays s ON p.patient_id = s.patient_id
        WHERE p.first_name LIKE ? COLLATE NOCASE
           OR p.last_name LIKE ? COLLATE NOCASE
           OR s.primary_diagnosis LIKE ? COLLATE NOCASE
        ORDER BY s.admission_date DESC
    """, (like, like, like)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_stay_detail(stay_id):
    conn = _connect()
    row = conn.execute("""
        SELECT p.first_name, p.last_name, p.mrn, p.dob,
               s.*
        FROM hospital_stays s
        JOIN patients p ON p.patient_id = s.patient_id
        WHERE s.stay_id = ?
    """, (stay_id,)).fetchone()
    conn.close()
    return dict(row) if row else None


def add_patient_and_stay(data):
    """data: dict with patient + stay fields, used by the 'Add Record' screen."""
    conn = _connect()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO patients (first_name, last_name, dob, mrn) VALUES (?, ?, ?, ?)",
        (data["first_name"], data["last_name"], data["dob"], data["mrn"]),
    )
    patient_id = cur.lastrowid
    cur.execute("""
        INSERT INTO hospital_stays
        (patient_id, primary_diagnosis, daily_progress_notes, hospital_course,
         treatment_given, discharge_condition, discharge_advice)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (
        patient_id, data["primary_diagnosis"], data.get("daily_progress_notes", ""),
        data.get("hospital_course", ""), data.get("treatment_given", ""),
        data.get("discharge_condition", ""), data.get("discharge_advice", ""),
    ))
    conn.commit()
    conn.close()


def export_rows_to_excel(rows, filename=None):
    """Export a list of dicts (from search_patients) to an .xlsx file. Returns file path."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Patient Records"

    headers = ["Patient ID", "First Name", "Last Name", "MRN", "DOB",
               "Stay ID", "Diagnosis", "Admission Date", "Discharge Date"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2E5D9F", end_color="2E5D9F", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for r in rows:
        ws.append([
            r.get("patient_id"), r.get("first_name"), r.get("last_name"),
            r.get("mrn"), r.get("dob"), r.get("stay_id"),
            r.get("primary_diagnosis"), r.get("admission_date"), r.get("discharge_date"),
        ])

    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    if not filename:
        filename = f"patient_export_{date.today().isoformat()}.xlsx"
    path = os.path.join(get_export_dir(), filename)
    wb.save(path)
    return path
